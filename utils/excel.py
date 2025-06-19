# -*- coding: utf-8 -*-

import pandas as pd
import logging
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
from sqlalchemy.engine import Engine
from sqlalchemy import create_engine
import win32com.client
import pythoncom

logger = logging.getLogger(__name__)

class ExcelUtils:
    """
    Classe utilitaire pour gérer les opérations sur les fichiers Excel, y compris lecture, écriture, formules, conversion PDF, et interactions SQL.
    """

    @staticmethod
    def load_excel_to_dataframe(filepath: str, sheet_name: Union[str, int] = 0, **kwargs) -> Optional[pd.DataFrame]:
        """
        Charge une feuille spécifique d'un fichier Excel dans un DataFrame pandas.
        """
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, **kwargs)
            logger.info(f"Fichier Excel lu avec succès : {filepath}")
            return df
        except FileNotFoundError:
            logger.error(f"Erreur : Le fichier Excel '{filepath}' est introuvable.")
            return None
        except Exception as e:
            logger.error(f"Erreur lors du chargement du fichier Excel '{filepath}': {e}")
            return None

    @staticmethod
    def write_dataframe_to_excel(dataframe: pd.DataFrame, filepath: str, sheet_name: str = 'Sheet1', startrow: Optional[int] = None, startcol: Optional[int] = None, **kwargs):
        """
        Écrit un DataFrame pandas dans un fichier Excel, à l'emplacement et la feuille spécifiés.
        """
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a' if Path(filepath).exists() else 'w') as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow, startcol=startcol, **kwargs)
        logger.info(f"DataFrame sauvegardé en Excel : {filepath}")

    @staticmethod
    def write_to_excel_cell(filepath: str, sheet_name: str, row: int, col: int, value: Any):
        """
        Écrit une valeur dans une cellule spécifique d'un fichier Excel.
        """
        try:
            workbook = openpyxl.load_workbook(filepath)
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            sheet.cell(row=row, column=col, value=value)
            workbook.save(filepath)
            logger.info(f"Valeur écrite dans {sheet_name}!{row},{col}")
        except FileNotFoundError:
            logger.error(f"Erreur : Le fichier Excel '{filepath}' est introuvable.")
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture dans la cellule ({row}, {col}) de la feuille '{sheet_name}': {e}")

    @staticmethod
    def write_formula(filepath: str, sheet_name: str, cell: str, formula: str):
        """
        Écrit une formule dans une cellule Excel.
        """
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name]
            ws[cell] = formula
            wb.save(filepath)
            logger.info(f"Formule écrite dans {sheet_name}!{cell}")
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture de la formule: {str(e)}")

    @staticmethod
    def write_custom_function(filepath: str, sheet_name: str, cell: str, function_name: str, params: List[Any]):
        """
        Écrit une fonction personnalisée dans une cellule Excel.
        """
        try:
            params_str = ','.join(str(p) for p in params)
            formula = f"={function_name}({params_str})"
            ExcelUtils.write_formula(filepath, sheet_name, cell, formula)
            logger.info(f"Fonction personnalisée écrite dans {sheet_name}!{cell}")
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture de la fonction: {str(e)}")

    @staticmethod
    def clear_sheet(filepath: str, sheet_name: str):
        """
        Vide une feuille Excel.
        """
        try:
            wb = openpyxl.load_workbook(filepath)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        cell.value = None
                wb.save(filepath)
                logger.info(f"Feuille vidée : {sheet_name}")
        except Exception as e:
            logger.error(f"Erreur lors du vidage de la feuille: {str(e)}")

    @staticmethod
    def clear_range(filepath: str, sheet_name: str, cell_range: str):
        """
        Vide une plage spécifique dans une feuille Excel.
        """
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name]
            for row in ws[cell_range]:
                for cell in row:
                    cell.value = None
            wb.save(filepath)
            logger.info(f"Plage {cell_range} vidée dans {sheet_name}")
        except Exception as e:
            logger.error(f"Erreur lors du vidage de la plage: {str(e)}")

    @staticmethod
    def excel_to_dict(filepath: str, **kwargs) -> Dict[str, pd.DataFrame]:
        """
        Convertit un fichier Excel en dictionnaire de DataFrames (une par feuille).
        """
        try:
            excel_file = pd.ExcelFile(filepath)
            result = {
                sheet_name: pd.read_excel(excel_file, sheet_name=sheet_name, **kwargs)
                for sheet_name in excel_file.sheet_names
            }
            logger.info(f"Excel converti en dictionnaire : {filepath}")
            return result
        except Exception as e:
            logger.error(f"Erreur lors de la conversion Excel->Dict: {str(e)}")
            return {}

    @staticmethod
    def excel_to_sql(df: pd.DataFrame, table_name: str, connection: Engine, if_exists: str = 'fail', **kwargs):
        """
        Importe un DataFrame (issu d'Excel) dans une table SQL.
        """
        df.to_sql(table_name, connection, if_exists=if_exists, index=False, **kwargs)
        logger.info(f"DataFrame importé dans SQL : {table_name}")

    @staticmethod
    def excel_file_to_sql(excel_path: str, table_name: str, connection: Engine, sheet_name: Union[str, int] = 0, if_exists: str = 'fail', **kwargs):
        """
        Importe une feuille Excel dans une table SQL.
        """
        df = ExcelUtils.load_excel_to_dataframe(excel_path, sheet_name=sheet_name)
        ExcelUtils.excel_to_sql(df, table_name, connection, if_exists=if_exists, **kwargs)

    @staticmethod
    def sql_to_excel(query: str, excel_path: str, connection: Engine, **kwargs):
        """
        Exporte le résultat d'une requête SQL vers un fichier Excel.
        """
        df = pd.read_sql_query(query, connection)
        ExcelUtils.write_dataframe_to_excel(df, excel_path, **kwargs)
        logger.info(f"Données SQL exportées en Excel : {excel_path}")

    @staticmethod
    def excel_to_pdf(excel_path: str, pdf_path: str):
        """
        Convertit un fichier Excel en PDF (Windows/Excel requis).
        """
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(excel_path)
            wb.ExportAsFixedFormat(0, pdf_path)
            wb.Close()
            excel.Quit()
            logger.info(f"Excel converti en PDF : {pdf_path}")
        except Exception as e:
            logger.error(f"Erreur lors de la conversion Excel->PDF: {str(e)}")
        finally:
            pythoncom.CoUninitialize()
