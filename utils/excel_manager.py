"""
Module de gestion des fichiers Excel.
"""

import logging
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List, Union, Optional, Any, Tuple
from pathlib import Path
import win32com.client
import pythoncom
from sqlalchemy import create_engine

logger = logging.getLogger(__name__)

class ExcelManager:
    """Gestionnaire de fichiers Excel."""
    
    def __init__(self):
        """Initialise le gestionnaire Excel."""
        self.workbook = None
        self.file_path = None
    
    def open_workbook(self, file_path: Union[str, Path]) -> None:
        """
        Ouvre un classeur Excel.
        
        Args:
            file_path: Chemin du fichier Excel
        """
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self.file_path = file_path
            logger.info(f"Classeur Excel ouvert: {file_path}")
        except Exception as e:
            logger.error(f"Erreur lors de l'ouverture du classeur {file_path}: {str(e)}")
            raise
    
    def create_workbook(self) -> None:
        """Crée un nouveau classeur Excel."""
        self.workbook = openpyxl.Workbook()
        logger.info("Nouveau classeur Excel créé")
    
    def read_sheet(
        self,
        sheet_name: Optional[str] = None,
        usecols: Optional[List[str]] = None,
        **kwargs
    ) -> pd.DataFrame:
        """
        Lit une feuille Excel en DataFrame.
        
        Args:
            sheet_name: Nom de la feuille
            usecols: Colonnes à lire
            **kwargs: Arguments additionnels pour pd.read_excel
            
        Returns:
            DataFrame: Données de la feuille
        """
        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                usecols=usecols,
                **kwargs
            )
            logger.info(f"Feuille Excel lue: {sheet_name or 'default'}")
            return df
        except Exception as e:
            logger.error(f"Erreur lors de la lecture de la feuille: {str(e)}")
            raise
    
    def write_dataframe(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        start_cell: str = 'A1',
        **kwargs
    ) -> None:
        """
        Écrit un DataFrame dans une feuille Excel.
        
        Args:
            df: DataFrame à écrire
            sheet_name: Nom de la feuille
            start_cell: Cellule de départ
            **kwargs: Arguments additionnels
        """
        try:
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
            else:
                sheet = self.workbook.create_sheet(sheet_name)
            
            rows = dataframe_to_rows(df, index=False, header=True)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = sheet.cell(row=r_idx, column=c_idx)
                    cell.value = value
            
            logger.info(f"DataFrame écrit dans la feuille: {sheet_name}")
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture du DataFrame: {str(e)}")
            raise
    
    def write_formula(
        self,
        sheet_name: str,
        cell: str,
        formula: str
    ) -> None:
        """
        Écrit une formule dans une cellule.
        
        Args:
            sheet_name: Nom de la feuille
            cell: Référence de la cellule
            formula: Formule Excel
        """
        try:
            sheet = self.workbook[sheet_name]
            sheet[cell].value = f"={formula}"
            logger.info(f"Formule écrite dans {sheet_name}!{cell}")
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture de la formule: {str(e)}")
            raise
    
    def write_custom_function(
        self,
        sheet_name: str,
        cell: str,
        vba_code: str
    ) -> None:
        """
        Écrit une fonction VBA personnalisée.
        
        Args:
            sheet_name: Nom de la feuille
            cell: Référence de la cellule
            vba_code: Code VBA de la fonction
        """
        try:
            if not self.workbook.vba_archive:
                self.workbook.create_vba_module()
            
            module = self.workbook.vba_archive.vba_modules[0]
            module.code = vba_code
            
            sheet = self.workbook[sheet_name]
            sheet[cell].value = f"=CustomFunction()"
            logger.info(f"Fonction VBA écrite dans {sheet_name}!{cell}")
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture de la fonction VBA: {str(e)}")
            raise
    
    def save(self, file_path: Optional[Union[str, Path]] = None) -> None:
        """
        Sauvegarde le classeur Excel.
        
        Args:
            file_path: Nouveau chemin de fichier (optionnel)
        """
        try:
            save_path = file_path or self.file_path
            self.workbook.save(save_path)
            logger.info(f"Classeur sauvegardé: {save_path}")
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde: {str(e)}")
            raise
    
    def refresh_all(self) -> None:
        """Actualise toutes les données du classeur."""
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            wb = excel.Workbooks.Open(self.file_path)
            wb.RefreshAll()
            wb.Save()
            wb.Close()
            excel.Quit()
            logger.info("Données du classeur actualisées")
        except Exception as e:
            logger.error(f"Erreur lors de l'actualisation: {str(e)}")
            raise
        finally:
            pythoncom.CoUninitialize()
    
    def to_pdf(self, pdf_path: Union[str, Path]) -> None:
        """
        Convertit le classeur en PDF.
        
        Args:
            pdf_path: Chemin du fichier PDF
        """
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            wb = excel.Workbooks.Open(self.file_path)
            wb.ExportAsFixedFormat(0, pdf_path)  # 0 = PDF format
            wb.Close()
            excel.Quit()
            logger.info(f"Classeur converti en PDF: {pdf_path}")
        except Exception as e:
            logger.error(f"Erreur lors de la conversion en PDF: {str(e)}")
            raise
        finally:
            pythoncom.CoUninitialize()
    
    def to_sql_server(
        self,
        sheet_name: str,
        table_name: str,
        connection_string: str,
        schema: str = 'dbo',
        if_exists: str = 'fail'
    ) -> None:
        """
        Sauvegarde une feuille dans SQL Server.
        
        Args:
            sheet_name: Nom de la feuille
            table_name: Nom de la table
            connection_string: Chaîne de connexion SQL Server
            schema: Schéma de la base de données
            if_exists: Comportement si la table existe
        """
        try:
            df = self.read_sheet(sheet_name)
            engine = create_engine(f"mssql+pyodbc:///?odbc_connect={connection_string}")
            df.to_sql(
                table_name,
                engine,
                schema=schema,
                if_exists=if_exists,
                index=False
            )
            logger.info(f"Feuille sauvegardée dans SQL Server: {table_name}")
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde dans SQL Server: {str(e)}")
            raise
    
    def to_sqlite(
        self,
        sheet_name: str,
        table_name: str,
        db_path: Union[str, Path],
        if_exists: str = 'fail'
    ) -> None:
        """
        Sauvegarde une feuille dans SQLite.
        
        Args:
            sheet_name: Nom de la feuille
            table_name: Nom de la table
            db_path: Chemin de la base SQLite
            if_exists: Comportement si la table existe
        """
        try:
            df = self.read_sheet(sheet_name)
            engine = create_engine(f"sqlite:///{db_path}")
            df.to_sql(
                table_name,
                engine,
                if_exists=if_exists,
                index=False
            )
            logger.info(f"Feuille sauvegardée dans SQLite: {table_name}")
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde dans SQLite: {str(e)}")
            raise
    
    def to_dict(self, sheet_name: str) -> Dict[str, Any]:
        """
        Convertit une feuille en dictionnaire.
        
        Args:
            sheet_name: Nom de la feuille
            
        Returns:
            Dict: Données de la feuille
        """
        try:
            df = self.read_sheet(sheet_name)
            return df.to_dict('records')
        except Exception as e:
            logger.error(f"Erreur lors de la conversion en dictionnaire: {str(e)}")
            raise 