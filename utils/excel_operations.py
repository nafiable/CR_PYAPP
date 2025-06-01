"""
Module de gestion des opérations sur les fichiers Excel.
"""

import logging
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
from database.connexionsqlServer import SQLServerConnection
from database.connexionsqlLiter import SQLiteConnection

logger = logging.getLogger(__name__)

class ExcelOperations:
    """Classe pour gérer les opérations sur les fichiers Excel."""
    
    def __init__(self):
        """Initialisation des connexions aux bases de données."""
        self.sql_server = SQLServerConnection()
        self.sqlite = SQLiteConnection()
    
    def read_excel(self, file_path: str, **kwargs) -> pd.DataFrame:
        """
        Lit un fichier Excel et retourne un DataFrame.
        
        Args:
            file_path (str): Chemin du fichier Excel
            **kwargs: Arguments supplémentaires pour pd.read_excel
            
        Returns:
            pd.DataFrame: DataFrame contenant les données
        """
        try:
            df = pd.read_excel(file_path, **kwargs)
            logger.info(f"Fichier Excel lu avec succès : {file_path}")
            return df
        except Exception as e:
            logger.error(f"Erreur lors de la lecture du fichier Excel {file_path}: {str(e)}")
            raise
    
    def write_excel(self, df: pd.DataFrame, file_path: str, **kwargs) -> None:
        """
        Écrit un DataFrame dans un fichier Excel.
        
        Args:
            df (pd.DataFrame): DataFrame à sauvegarder
            file_path (str): Chemin du fichier Excel
            **kwargs: Arguments supplémentaires pour df.to_excel
        """
        try:
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            df.to_excel(file_path, **kwargs)
            logger.info(f"DataFrame sauvegardé en Excel : {file_path}")
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture du fichier Excel {file_path}: {str(e)}")
            raise
    
    def excel_to_sql_server(self, excel_path: str, table_name: str, if_exists: str = 'fail', **kwargs) -> None:
        """
        Importe un fichier Excel dans SQL Server.
        
        Args:
            excel_path (str): Chemin du fichier Excel
            table_name (str): Nom de la table
            if_exists (str): Comportement si la table existe ('fail', 'replace', 'append')
            **kwargs: Arguments supplémentaires pour pd.read_excel
        """
        try:
            df = self.read_excel(excel_path, **kwargs)
            with self.sql_server.get_connection() as conn:
                df.to_sql(table_name, conn, if_exists=if_exists, index=False)
            logger.info(f"Excel importé dans SQL Server : {table_name}")
        except Exception as e:
            logger.error(f"Erreur lors de l'import Excel->SQL Server: {str(e)}")
            raise
    
    def excel_to_sqlite(self, excel_path: str, table_name: str, if_exists: str = 'fail', **kwargs) -> None:
        """
        Importe un fichier Excel dans SQLite.
        
        Args:
            excel_path (str): Chemin du fichier Excel
            table_name (str): Nom de la table
            if_exists (str): Comportement si la table existe ('fail', 'replace', 'append')
            **kwargs: Arguments supplémentaires pour pd.read_excel
        """
        try:
            df = self.read_excel(excel_path, **kwargs)
            with self.sqlite.get_connection() as conn:
                df.to_sql(table_name, conn, if_exists=if_exists, index=False)
            logger.info(f"Excel importé dans SQLite : {table_name}")
        except Exception as e:
            logger.error(f"Erreur lors de l'import Excel->SQLite: {str(e)}")
            raise
    
    def sql_to_excel(self, query: str, excel_path: str, is_sqlite: bool = False, **kwargs) -> None:
        """
        Exporte le résultat d'une requête SQL vers un fichier Excel.
        
        Args:
            query (str): Requête SQL
            excel_path (str): Chemin du fichier Excel
            is_sqlite (bool): True pour SQLite, False pour SQL Server
            **kwargs: Arguments supplémentaires pour df.to_excel
        """
        try:
            conn = self.sqlite if is_sqlite else self.sql_server
            with conn.get_connection() as connection:
                df = pd.read_sql_query(query, connection)
                self.write_excel(df, excel_path, **kwargs)
            logger.info(f"Données SQL exportées en Excel : {excel_path}")
        except Exception as e:
            logger.error(f"Erreur lors de l'export SQL->Excel: {str(e)}")
            raise
    
    def write_formula(self, file_path: str, sheet_name: str, cell: str, formula: str) -> None:
        """
        Écrit une formule dans une cellule Excel.
        
        Args:
            file_path (str): Chemin du fichier Excel
            sheet_name (str): Nom de la feuille
            cell (str): Référence de la cellule (ex: 'A1')
            formula (str): Formule Excel (commençant par '=')
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name]
            ws[cell] = formula
            wb.save(file_path)
            logger.info(f"Formule écrite dans {sheet_name}!{cell}")
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture de la formule: {str(e)}")
            raise
    
    def write_custom_function(self, file_path: str, sheet_name: str, cell: str, 
                            function_name: str, params: List[Any]) -> None:
        """
        Écrit une fonction personnalisée dans une cellule Excel.
        
        Args:
            file_path (str): Chemin du fichier Excel
            sheet_name (str): Nom de la feuille
            cell (str): Référence de la cellule
            function_name (str): Nom de la fonction
            params (List[Any]): Liste des paramètres
        """
        try:
            # Convertir les paramètres en chaîne de caractères
            params_str = ','.join(str(p) for p in params)
            formula = f"={function_name}({params_str})"
            
            self.write_formula(file_path, sheet_name, cell, formula)
            logger.info(f"Fonction personnalisée écrite dans {sheet_name}!{cell}")
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture de la fonction: {str(e)}")
            raise
    
    def refresh_workbook(self, file_path: str) -> None:
        """
        Rafraîchit toutes les formules d'un classeur Excel.
        
        Args:
            file_path (str): Chemin du fichier Excel
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            cell.value = cell.value  # Recalcule la formule
            wb.save(file_path)
            logger.info(f"Classeur rafraîchi : {file_path}")
        except Exception as e:
            logger.error(f"Erreur lors du rafraîchissement du classeur: {str(e)}")
            raise
    
    def excel_to_pdf(self, excel_path: str, pdf_path: str) -> None:
        """
        Convertit un fichier Excel en PDF.
        
        Args:
            excel_path (str): Chemin du fichier Excel
            pdf_path (str): Chemin du fichier PDF
        """
        try:
            import win32com.client
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            
            wb = excel.Workbooks.Open(excel_path)
            wb.ExportAsFixedFormat(0, pdf_path)  # 0 = PDF format
            
            wb.Close()
            excel.Quit()
            
            logger.info(f"Excel converti en PDF : {pdf_path}")
        except Exception as e:
            logger.error(f"Erreur lors de la conversion Excel->PDF: {str(e)}")
            raise
    
    @staticmethod
    def df_to_excel(df: pd.DataFrame, excel_path: str, **kwargs) -> None:
        """
        Convertit un DataFrame en fichier Excel.
        
        Args:
            df (pd.DataFrame): DataFrame à convertir
            excel_path (str): Chemin du fichier Excel
            **kwargs: Arguments supplémentaires pour df.to_excel
        """
        try:
            Path(excel_path).parent.mkdir(parents=True, exist_ok=True)
            df.to_excel(excel_path, **kwargs)
            logger.info(f"DataFrame converti en Excel : {excel_path}")
        except Exception as e:
            logger.error(f"Erreur lors de la conversion DataFrame->Excel: {str(e)}")
            raise
    
    @staticmethod
    def excel_to_dict(excel_path: str, **kwargs) -> Dict[str, pd.DataFrame]:
        """
        Convertit un fichier Excel en dictionnaire de DataFrames.
        
        Args:
            excel_path (str): Chemin du fichier Excel
            **kwargs: Arguments supplémentaires pour pd.read_excel
            
        Returns:
            Dict[str, pd.DataFrame]: Dictionnaire {nom_feuille: dataframe}
        """
        try:
            excel_file = pd.ExcelFile(excel_path)
            result = {
                sheet_name: pd.read_excel(excel_file, sheet_name=sheet_name, **kwargs)
                for sheet_name in excel_file.sheet_names
            }
            logger.info(f"Excel converti en dictionnaire : {excel_path}")
            return result
        except Exception as e:
            logger.error(f"Erreur lors de la conversion Excel->Dict: {str(e)}")
            raise 